from itertools import combinations

from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties, WorksheetProperties
from openpyxl.worksheet.worksheet import Worksheet

from campx.model.camp import Camp
from campx.model.enums import EntryType


def fill_non_overlapping_pairs_sheet(camp: Camp, ws: Worksheet) -> None:
    """List leader pairs that never overlap in day-off and sleep-in assignments."""
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
    ws.print_options.gridLines = True
    ws.print_options.horizontalCentered = True
    ws.sheet_properties = WorksheetProperties(
        pageSetUpPr=PageSetupProperties(fitToPage=True)
    )

    leaders = camp.leaders_incl_management
    day_off_by_leader = {leader.participant_id: set() for leader in leaders}
    sleep_in_by_leader = {leader.participant_id: set() for leader in leaders}

    for day in camp.schedule.days:
        for entry in day.schedule_entries:
            if entry.entry_type not in (EntryType.DAY_OFF, EntryType.SLEEP_IN):
                continue
            days_by_leader = (
                day_off_by_leader
                if entry.entry_type == EntryType.DAY_OFF
                else sleep_in_by_leader
            )
            for responsible in entry.responsible:
                if responsible.participant_id in days_by_leader:
                    days_by_leader[responsible.participant_id].add(day.date)

    ws.cell(row=1, column=1, value="Leader 1").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Leader 2").font = Font(bold=True)

    row = 2
    for leader_a, leader_b in combinations(leaders, 2):
        leader_a_day_off = day_off_by_leader[leader_a.participant_id]
        leader_b_day_off = day_off_by_leader[leader_b.participant_id]
        leader_a_sleep_in = sleep_in_by_leader[leader_a.participant_id]
        leader_b_sleep_in = sleep_in_by_leader[leader_b.participant_id]

        same_day_off = leader_a_day_off & leader_b_day_off
        same_sleep_in = leader_a_sleep_in & leader_b_sleep_in
        cross_overlap = (leader_a_day_off & leader_b_sleep_in) | (
            leader_a_sleep_in & leader_b_day_off
        )

        if same_day_off or same_sleep_in or cross_overlap:
            continue

        ws.cell(row=row, column=1, value=leader_a.full_name)
        ws.cell(row=row, column=2, value=leader_b.full_name)
        row += 1
        ws.cell(row=row, column=1, value=leader_b.full_name)
        ws.cell(row=row, column=2, value=leader_a.full_name)
        row += 1

    if row == 2:
        ws.cell(row=row, column=1, value="No non-overlapping pairs found.")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    for col in range(1, ws.max_column + 1):
        max_len = max(
            (
                len(str(ws.cell(row=r, column=col).value or ""))
                for r in range(1, ws.max_row + 1)
            ),
            default=0,
        )
        ws.column_dimensions[get_column_letter(col)].width = max(18, max_len + 2)

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).alignment = Alignment(
                vertical="center", horizontal="left"
            )
