from campx.model.camp import Camp
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties


def fill_working_hours_sheet(camp: Camp, ws: Worksheet):
    """Fill a sheet showing working hours per leader per day (based on timed activities only)."""
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToPage = True

    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
    ws.print_options.gridLines = True
    ws.print_options.horizontalCentered = True
    ws.sheet_properties = WorksheetProperties(
        pageSetUpPr=PageSetupProperties(fitToPage=True)
    )

    # Header row with day columns
    ws.cell(row=1, column=1, value="Namn")
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(
        vertical="center", horizontal="center"
    )

    for col, day in enumerate(camp.schedule.days, start=2):
        day_header = day.as_str(
            "%a %d %b",
            swedish_month_names=True,
            swedish_weekday_names=True,
        )
        cell = ws.cell(row=1, column=col, value=day_header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", horizontal="center")

    # Data rows: one per leader
    for row, leader in enumerate(camp.leaders_incl_management, start=2):
        ws.cell(row=row, column=1, value=leader.full_name)
        ws.cell(row=row, column=1).font = Font(bold=True)

        for col, day in enumerate(camp.schedule.days, start=2):
            # Get all timed entries for this leader on this day
            start_times = []
            end_times = []
            for entry in day.schedule_entries:
                # Only count timed activities (those with start_time and end_time, not time_independent)
                if (
                    entry.start_time
                    and entry.end_time
                    and not entry.entry_type.value.time_independent
                    and leader in entry.responsible
                ):
                    start_times.append(entry.start_time)
                    end_times.append(entry.end_time)

            # Display time range if leader has timed entries
            if start_times and end_times:
                earliest_start = min(start_times)
                latest_end = max(end_times)
                time_range = f"{earliest_start} - {latest_end}"
                cell = ws.cell(row=row, column=col, value=time_range)
                cell.alignment = Alignment(vertical="center", horizontal="center")
                # Light background for cells with hours
                cell.fill = PatternFill(
                    start_color="E8F4F8", end_color="E8F4F8", fill_type="solid"
                )

    # Auto-adjust column widths
    for col_idx in range(1, ws.max_column + 1):
        max_len = max(
            (
                len(str(ws.cell(row=row, column=col_idx).value or ""))
                for row in range(1, ws.max_row + 1)
            ),
            default=0,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, max_len + 2)
