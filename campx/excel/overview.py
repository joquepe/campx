from campx.model.camp import Camp
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from campx.model.enums import ParticipantType, EntryType
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from campx.excel.utilities import get_activity_color
from campx.excel.utilities import Colors


def get_fill_color(value: int, max_value: int = 10) -> PatternFill:
    if value == 0:
        return None  # No fill

    # Clamp value between 1 and max_value
    value = max(1, min(value, max_value))

    # Create a gradient from green (value=1) to red (value=max_value)
    red = int((value - 1) / (max_value - 1) * 255)
    green = int((max_value - value) / (max_value - 1) * 255)

    hex_color = f"{red:02X}{green:02X}00"
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def fill_overview_sheet(camp: Camp, ws: Worksheet):
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = "landscape"  # Optional
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 0  # Use 1 if you also want to fit to 1 page wide

    # Optional: Set margins and scaling if needed
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

    # Optional: Print gridlines if you want the cell lines visible
    ws.print_options.gridLines = True
    ws.print_options.horizontalCentered = True

    # Optional: Make sure worksheet is printable
    ws.sheet_properties = WorksheetProperties(
        pageSetUpPr=PageSetupProperties(fitToPage=True)
    )

    ws.cell(row=1, column=1, value="Namn")
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=1).border = Border(
        bottom=Side(style="medium", color=Colors.BLACK),
        right=Side(style="thin", color=Colors.BLACK),
    )
    ws.cell(row=1, column=1).alignment = Alignment(
        vertical="center", horizontal="center"
    )

    col = 2
    for entry_type in EntryType:
        cell = ws.cell(row=1, column=col, value=entry_type.value.short)
        cell.font = Font(italic=True, bold=True)
        color = get_activity_color(entry_type)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(vertical="center", horizontal="center")
        col += 1

    row = 2
    for leader in camp.leaders_incl_management:
        ws.cell(row=row, column=1, value=leader.full_name)
        if leader.participant_type == ParticipantType.CAMP_MANAGEMENT:
            ws.cell(row=row, column=1).font = Font(
                bold=True, italic=True, color=Colors.GREY
            )
        else:
            ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).border = Border(
            bottom=Side(style="thin", color=Colors.GREY),
            right=Side(style="thin", color=Colors.BLACK),
        )

        col = 2
        for entry_type in EntryType:
            entries = leader.get_responsible_entries(
                camp.schedule, entry_type=entry_type
            )
            num_entries = len(entries)
            cell = ws.cell(row=row, column=col, value=num_entries)
            cell.border = Border(
                bottom=Side(style="thin", color=Colors.GREY),
                right=Side(style="thin", color=Colors.BLACK),
            )
            fill = get_fill_color(num_entries)
            if fill:
                cell.fill = fill
            col += 1

        row += 1

    for col in range(1, ws.max_column + 1):
        max_len = max(
            (
                len(str(ws.cell(row=r, column=col).value or ""))
                for r in range(1, ws.max_row + 1)
            ),
            default=0,
        )
        ws.column_dimensions[get_column_letter(col)].width = max(10, max_len + 2)
