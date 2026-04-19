from campx.model.camp import Camp
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.utils import get_column_letter
from campx.model.schedule_entry import ScheduleEntry
from campx.model.enums import EntryType
from campx.model.participant import Participant
from campx.model.schedule import Schedule
from campx.excel.utilities import START_HOUR, END_HOUR, get_activity_color
from campx.excel.utilities import Colors

INVALID_SHEET_CHARS = set("[]:*?/\\")
MAX_SHEET_NAME_LENGTH = 31
DAY_COLUMN_WIDTH = 18
TIME_SLOT_ROW_HEIGHT = 15
TIME_SLOT_COLUMN_WIDTH = 12


def _sanitize_sheet_name(name: str) -> str:
    cleaned = "".join(ch for ch in name if ch not in INVALID_SHEET_CHARS).strip()
    if not cleaned:
        cleaned = "Participant"
    if len(cleaned) > MAX_SHEET_NAME_LENGTH:
        cleaned = cleaned[:MAX_SHEET_NAME_LENGTH]
    return cleaned


def _make_unique_sheet_name(workbook: Workbook, base_name: str) -> str:
    sanitized = _sanitize_sheet_name(base_name)
    if sanitized not in workbook.sheetnames:
        return sanitized

    suffix = 1
    while True:
        suffix_name = (
            f"{sanitized[: MAX_SHEET_NAME_LENGTH - len(str(suffix)) - 1]}_{suffix}"
        )
        if suffix_name not in workbook.sheetnames:
            return suffix_name
        suffix += 1


def _build_schedule_entry_value(
    header_text: str,
    responsible_text: str = "",
    additional_info: str = "",
) -> str:
    lines = []
    if header_text:
        lines.append(header_text)
    if responsible_text:
        lines.append(responsible_text)
    if additional_info:
        lines.append(additional_info)
    return "\n".join(lines)


def _get_secondary_row(row_num: int) -> int:
    return row_num + 1


def fill_schedule_sheet(
    camp: Camp, worksheet: Worksheet, highlight_tokens: set[str] | None = None
):
    worksheet.page_setup.paperSize = 8  # A3
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToHeight = 1
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToPage = True
    worksheet.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
    worksheet.print_options.gridLines = True
    worksheet.print_options.horizontalCentered = True
    worksheet.sheet_properties = WorksheetProperties(
        pageSetUpPr=PageSetupProperties(fitToPage=True)
    )
    top_types = EntryType.time_independent_top()
    bottom_types = EntryType.time_independent_bottom()
    buffer_rows_top = len(top_types)
    time_slots = get_time_slots(START_HOUR, END_HOUR, buffer_rows_top)

    # Create first column with hours
    for slot, row_num in time_slots.items():
        worksheet.cell(row=row_num, column=1, value=slot)

    # Create time-independent rows above the date row
    for row_offset, entry_type in enumerate(top_types, start=1):
        worksheet.cell(row=row_offset, column=1, value=entry_type.value.long)

    # Create date headers
    date_row = buffer_rows_top + 1
    for col, day in enumerate(camp.schedule.days, start=2):
        date_cell = worksheet.cell(row=date_row, column=col, value=day.as_str("%d %b"))
        date_cell.font = Font(bold=True)

    # Fill top time-independent values per day
    for col, day in enumerate(camp.schedule.days, start=2):
        for row_offset, entry_type in enumerate(top_types, start=1):
            entries = day.get_entries(entry_type)
            if not entries:
                continue
            values = []
            for entry in entries:
                responsible_text = ", ".join(
                    leader.nick_name or leader.initials for leader in entry.responsible
                )
                values.append(
                    _build_schedule_entry_value(
                        entry.name or "",
                        responsible_text,
                        entry.additional_info,
                    )
                )
            cell = worksheet.cell(
                row=row_offset,
                column=col,
                value="\n".join(v for v in values if v),
            )
            if entry_type == EntryType.THEME or any(entry.name for entry in entries):
                cell.font = Font(
                    bold=True,
                    italic=any(entry.responsible for entry in entries),
                )
            elif any(entry.responsible for entry in entries):
                cell.font = Font(italic=True)

    # Loop through each day in the schedule and add timed activities
    for col, day in enumerate(camp.schedule.days, start=2):
        timed_entries = [
            entry
            for entry in day.schedule_entries
            if not entry.entry_type.value.time_independent
            and entry.entry_type.value.show_in_schedule
        ]
        timed_entries.sort(key=lambda entry: entry.start_time or "")

        for activity in timed_entries:
            cell, value_to_cell = find_cell_for_activity(
                time_slots, activity, worksheet, col
            )
            secondary_row = _get_secondary_row(cell.row)
            secondary_cell = worksheet.cell(row=secondary_row, column=col)
            responsible_text = ", ".join(
                leader.nick_name or leader.initials for leader in activity.responsible
            )
            header_text = value_to_cell.split("\n", 1)[0]
            if cell.value:
                cell.value += "\n" + header_text
            else:
                cell.value = header_text

            secondary_value = _build_schedule_entry_value(
                "",
                responsible_text,
                activity.additional_info,
            )
            if secondary_value:
                if secondary_cell.value:
                    secondary_cell.value += "\n" + secondary_value
                else:
                    secondary_cell.value = secondary_value

            if activity.name:
                cell.font = Font(bold=True)
            if responsible_text or activity.additional_info:
                secondary_cell.font = Font(italic=bool(responsible_text))

            color = get_activity_color(activity.entry_type)
            cell.fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
            secondary_cell.fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )

    # Merge empty cells with the most previous cell that has data
    merge_empty_cells_with_previous(camp, worksheet, date_row, time_slots)

    # Create bottom time-independent rows after timed slots
    bottom_start_row = max(time_slots.values()) + 2
    for row_offset, entry_type in enumerate(bottom_types, start=1):
        worksheet.cell(
            row=bottom_start_row + row_offset - 1,
            column=1,
            value=entry_type.value.short,
        )

    # Fill bottom time-independent values per day
    for col, day in enumerate(camp.schedule.days, start=2):
        for row_offset, entry_type in enumerate(bottom_types, start=1):
            entries = day.get_entries(entry_type)
            if not entries:
                continue
            values = []
            for entry in entries:
                responsible_text = ", ".join(
                    leader.nick_name or leader.initials for leader in entry.responsible
                )
                values.append(
                    _build_schedule_entry_value(
                        entry.name or "",
                        responsible_text,
                        entry.additional_info,
                    )
                )
            cell = worksheet.cell(
                row=bottom_start_row + row_offset - 1,
                column=col,
                value="\n".join(v for v in values if v),
            )
            if any(entry.name for entry in entries):
                cell.font = Font(
                    bold=True,
                    italic=any(entry.responsible for entry in entries),
                )
            elif any(entry.responsible for entry in entries):
                cell.font = Font(italic=True)

    merge_time_slot_cells(time_slots, worksheet)
    total_rows = max(time_slots.values()) + len(bottom_types)
    add_borders(
        date_row, camp.schedule, worksheet, time_slots, bottom_types, total_rows
    )
    add_alignment(
        date_row, camp.schedule, worksheet, time_slots, bottom_types, total_rows
    )
    apply_uniform_schedule_dimensions(worksheet, camp, time_slots)


def fill_schedule_sheet_for_participant(
    camp: Camp, worksheet: Worksheet, participant: Participant
):
    fill_schedule_sheet(camp, worksheet)
    participant_label = participant.nick_name or participant.full_name
    label_col = worksheet.max_column + 2
    label_cell = worksheet.cell(
        row=1, column=label_col, value=f"Participant: {participant_label}"
    )
    label_cell.font = Font(bold=True, color="FF0000")


def make_unique_schedule_sheet_name(workbook: Workbook, participant: Participant):
    return _make_unique_sheet_name(
        workbook, participant.nick_name or participant.full_name
    )


def get_time_slots(start_hour: int, end_hour: int, extra_rows: int) -> dict[str, int]:
    time_slots = {}
    counter = start_hour
    for hour in range(start_hour, end_hour):
        for minute in ["00", "30"]:
            time_slots[f"{hour:02d}:{minute}"] = counter - (start_hour - 2 - extra_rows)
            counter += 2
    return time_slots


def get_effective_cell(worksheet: Worksheet, row: int, col: int):
    cell = worksheet.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for merged_range in worksheet.merged_cells.ranges:
            if (
                merged_range.min_row <= row <= merged_range.max_row
                and merged_range.min_col <= col <= merged_range.max_col
            ):
                return worksheet.cell(
                    row=merged_range.min_row, column=merged_range.min_col
                )
    return cell


def find_cell_for_activity(
    time_slots: dict, schedule_entry: ScheduleEntry, worksheet: Worksheet, col: int
):
    time_keys = list(time_slots.keys())
    start_time = schedule_entry.start_time
    start_row = None

    for i, slot in enumerate(time_keys):
        if start_time >= slot and (
            i + 1 == len(time_keys) or start_time < time_keys[i + 1]
        ):
            start_row = time_slots[slot]
            cell = get_effective_cell(worksheet, start_row, col)
            # For time-dependent activities, use custom name or fall back to long name
            value_to_cell = schedule_entry.name or schedule_entry.entry_type.value.long
            if schedule_entry.start_time != slot:
                value_to_cell += " " + schedule_entry.start_time
            if schedule_entry.responsible:
                str_of_responsibles = ", ".join(
                    leader.nick_name or leader.initials
                    for leader in schedule_entry.responsible
                )
                value_to_cell += "\n" + str_of_responsibles
            if schedule_entry.additional_info:
                value_to_cell += "\n" + schedule_entry.additional_info
            return cell, value_to_cell

    # Fallback to first slot if no matching time found
    first_slot = time_keys[0]
    return get_effective_cell(worksheet, time_slots[first_slot], col), ""


def merge_time_slot_cells(time_slots: dict[str, int], worksheet: Worksheet):
    start_row = None
    for time_slot, row_num in time_slots.items():
        if time_slot.endswith("00"):
            start_row = row_num
        elif start_row is not None:
            worksheet.merge_cells(
                start_row=start_row,
                start_column=1,
                end_row=_get_secondary_row(row_num),
                end_column=1,
            )
            start_row = None


def merge_empty_cells_with_previous(
    camp: Camp, worksheet: Worksheet, date_row: int, time_slots: dict
):
    max_row = _get_secondary_row(max(time_slots.values()))
    for col in range(2, len(camp.schedule.days) + 2):
        row = date_row + 1
        while row <= max_row:
            cell = worksheet.cell(row=row, column=col)
            if cell.value:
                # Find consecutive empty cells after this non-empty cell
                start_merge = row + 1
                row += 1
                while row <= max_row and not worksheet.cell(row=row, column=col).value:
                    row += 1
                if start_merge < row:
                    # Merge the empty cells into the non-empty cell
                    worksheet.merge_cells(
                        start_row=start_merge - 1,
                        start_column=col,
                        end_row=row - 1,
                        end_column=col,
                    )
            else:
                row += 1


def add_borders(
    date_row: int,
    schedule: Schedule,
    worksheet: Worksheet,
    time_slots: dict,
    bottom_types: list,
    total_rows: int,
):
    black_thin_side = Side(border_style="thin", color=Colors.BLACK)
    grey_thin_side = Side(border_style="thin", color="D3D3D3")
    for col in range(1, len(schedule.days) + 2):
        left = black_thin_side
        right = None
        if col == len(schedule.days) + 1:
            right = black_thin_side
        for row in range(1, total_rows + 1):
            cell = worksheet.cell(row, col)
            top = None
            bottom = None

            if row < date_row or col == 1:
                bottom = grey_thin_side
            if row == 1:
                top = black_thin_side
            if row == date_row or row == total_rows:
                bottom = black_thin_side

            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def add_alignment(
    date_row: int,
    schedule: Schedule,
    worksheet: Worksheet,
    time_slots: dict,
    bottom_types: list,
    total_rows: int,
):
    timed_end_row = _get_secondary_row(max(time_slots.values()))
    for col in range(1, len(schedule.days) + 2):
        for row in range(1, total_rows + 1):
            wrap_text = True
            if (col == 1 and row > date_row and row <= timed_end_row) or (
                row == date_row
            ):
                vertical = "center"
                horizontal = "center"
            else:
                vertical = "top"
                horizontal = "left"
            cell = worksheet.cell(row, col)
            if cell.value and "\n" in str(cell.value):
                wrap_text = True
            cell.alignment = Alignment(
                wrap_text=wrap_text, vertical=vertical, horizontal=horizontal
            )


def add_fonts(
    date_row: int,
    schedule: Schedule,
    worksheet: Worksheet,
    time_slots: dict,
    bottom_types: list,
    total_rows: int,
):
    for col in range(1, len(schedule.days) + 2):
        for row in range(1, total_rows + 1):
            bold = False
            italic = True

            cell = worksheet.cell(row, col)
            if row == 1 or row == date_row:
                bold = True
                italic = False

            cell.font = Font(bold=bold, italic=italic)


def apply_uniform_schedule_dimensions(
    worksheet: Worksheet, camp: Camp, time_slots: dict[str, int]
):
    for col in range(2, len(camp.schedule.days) + 2):
        worksheet.column_dimensions[get_column_letter(col)].width = DAY_COLUMN_WIDTH

    for row_num in time_slots.values():
        worksheet.row_dimensions[row_num].height = TIME_SLOT_ROW_HEIGHT
        worksheet.row_dimensions[
            _get_secondary_row(row_num)
        ].height = TIME_SLOT_ROW_HEIGHT

    worksheet.column_dimensions["A"].width = TIME_SLOT_COLUMN_WIDTH
