from collections import defaultdict
import datetime as dt
from itertools import combinations
import statistics
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties, WorksheetProperties
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from campx.model.camp import Camp
from campx.model.enums import EntryType, LeaderYearType, ParticipantType
from campx.validation import is_eligible_participant
from campx.excel.utilities import Colors


def _setup_worksheet(ws: Worksheet):
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
    ws.print_options.gridLines = True
    ws.print_options.horizontalCentered = True
    ws.sheet_properties = WorksheetProperties(
        pageSetUpPr=PageSetupProperties(fitToPage=True)
    )


def _max_consecutive_true(values: list[bool]) -> int:
    max_streak = 0
    current_streak = 0
    for value in values:
        if value:
            current_streak += 1
            max_streak = max(max_streak, current_streak)
        else:
            current_streak = 0
    return max_streak


def _safe_stdev(values: list[int]) -> float:
    if len(values) < 2:
        return 0.0
    return statistics.pstdev(values)


def _time_to_datetime(day_date: dt.date, time_value: str) -> dt.datetime:
    hour, minute = map(int, time_value.split(":"))
    return dt.datetime.combine(day_date, dt.time(hour=hour, minute=minute))


def _age_on_date(birthday: dt.date, reference_date: dt.date) -> int:
    age = reference_date.year - birthday.year
    if (reference_date.month, reference_date.day) < (birthday.month, birthday.day):
        age -= 1
    return age


def _default_spacing_entry_types() -> set[EntryType]:
    return {
        EntryType.WAKE_UP,
        EntryType.MORNING_PRAYER,
        EntryType.EDUCATION,
        EntryType.AFTERNOON_ACTIVITY,
        EntryType.FEEDBACK,
        EntryType.EVENING_ACTIVITY,
        EntryType.EVENING_PRAYER,
        EntryType.PUT_TO_BED,
    }


def _calculate_responsibility_spacing(
    camp: Camp,
    included_entry_types: set[EntryType] | None = None,
    tight_gap_minutes: int = 120,
) -> dict:
    included_types = included_entry_types or _default_spacing_entry_types()
    leader_spacing = []
    all_gap_minutes = []

    for leader in camp.leaders_incl_management:
        leader_entries = []
        for day in camp.schedule.days:
            for entry in day.schedule_entries:
                if entry.entry_type not in included_types:
                    continue
                if leader not in entry.responsible:
                    continue
                start = _time_to_datetime(day.date, entry.start_time)
                end = _time_to_datetime(day.date, entry.end_time)
                leader_entries.append((start, end))

        leader_entries.sort(key=lambda item: item[0])
        gaps = []
        for (_, current_end), (next_start, _) in zip(
            leader_entries, leader_entries[1:]
        ):
            gap_minutes = int((next_start - current_end).total_seconds() // 60)
            gaps.append(max(gap_minutes, 0))

        all_gap_minutes.extend(gaps)
        tight_gap_count = sum(1 for gap in gaps if gap <= tight_gap_minutes)
        gap_count = len(gaps)
        tight_ratio = (tight_gap_count / gap_count) if gap_count else 0.0
        avg_gap_hours = (statistics.mean(gaps) / 60) if gaps else 0.0

        leader_spacing.append(
            {
                "leader_name": leader.full_name,
                "gap_count": gap_count,
                "tight_gap_count": tight_gap_count,
                "tight_gap_ratio": tight_ratio,
                "avg_gap_hours": avg_gap_hours,
            }
        )

    return {
        "included_entry_types": included_types,
        "tight_gap_minutes": tight_gap_minutes,
        "leader_spacing": leader_spacing,
        "all_gap_minutes": all_gap_minutes,
    }


def _write_table_headers(
    ws: Worksheet, start_row: int, start_col: int, headers: list[str]
):
    for offset, header in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + offset, value=header)
        cell.font = Font(bold=True)
        cell.border = Border(
            bottom=Side(style="thin", color=Colors.BLACK),
            right=Side(style="thin", color=Colors.BLACK),
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_adjust_column_widths(ws: Worksheet):
    for col_idx in range(1, ws.max_column + 1):
        max_len = max(
            (
                len(str(ws.cell(row=row, column=col_idx).value or ""))
                for row in range(1, ws.max_row + 1)
            ),
            default=0,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, max_len + 2)


def _calculate_leader_load(camp: Camp) -> list[dict]:
    loads = []
    for leader in camp.leaders_incl_management:
        responsibilities = leader.get_responsible_entries(camp.schedule)
        total = len(responsibilities)
        per_day = [
            len(leader.get_responsible_entries(camp.schedule, day))
            for day in camp.schedule.days
        ]
        active_days = sum(1 for count in per_day if count > 0)
        max_tasks_on_day = max(per_day) if per_day else 0
        max_consecutive = _max_consecutive_true([count > 0 for count in per_day])
        day_offs = len(
            leader.get_days_with_entry_type(EntryType.DAY_OFF, camp.schedule)
        )
        loads.append(
            {
                "leader": leader,
                "leader_name": leader.full_name,
                "total": total,
                "active_days": active_days,
                "max_tasks_on_day": max_tasks_on_day,
                "max_consecutive_active_days": max_consecutive,
                "day_offs": day_offs,
            }
        )
    return loads


def _calculate_eligibility(camp: Camp) -> list[dict]:
    eligibility = []
    for day in camp.schedule.days:
        for entry in day.schedule_entries:
            if not entry.entry_type.value.requires_responsible:
                continue
            eligible_count = 0
            for leader in camp.leaders_incl_management:
                if leader in entry.responsible:
                    continue
                is_eligible, _ = is_eligible_participant(leader, entry, camp, day)
                if is_eligible:
                    eligible_count += 1
            eligibility.append(
                {
                    "day": day,
                    "entry": entry,
                    "eligible_count": eligible_count,
                    "responsible_count": len(entry.responsible),
                }
            )
    return eligibility


def _calculate_interactions(camp: Camp) -> dict:
    pair_counts = defaultdict(int)
    unique_partner_sets = defaultdict(set)

    for day in camp.schedule.days:
        for entry in day.schedule_entries:
            responsible = entry.responsible
            for leader in responsible:
                for partner in responsible:
                    if partner is not leader:
                        unique_partner_sets[leader.full_name].add(partner.full_name)
            for leader1, leader2 in combinations(
                sorted(responsible, key=lambda p: p.full_name), 2
            ):
                pair_counts[(leader1.full_name, leader2.full_name)] += 1

    return {
        "pair_counts": pair_counts,
        "unique_partner_counts": {
            name: len(partners) for name, partners in unique_partner_sets.items()
        },
    }


def _calculate_entry_assignment_metrics(camp: Camp) -> list[dict]:
    entry_metrics = []

    for day in camp.schedule.days:
        for entry in day.schedule_entries:
            if not entry.entry_type.value.requires_responsible:
                continue

            responsibles = entry.responsible
            responsible_count = len(responsibles)
            first_year_count = sum(
                leader.leader_year_type == LeaderYearType.FIRST_YEAR
                for leader in responsibles
            )
            returning_leader_count = sum(
                leader.leader_year_type not in {None, LeaderYearType.FIRST_YEAR}
                for leader in responsibles
            )
            management_count = sum(
                leader.participant_type == ParticipantType.CAMP_MANAGEMENT
                for leader in responsibles
            )

            gender_counts = {
                gender: sum(leader.gender == gender for leader in responsibles)
                for gender in ("F", "M", "O")
            }
            gender_mix = (
                ", ".join(
                    f"{gender}:{count}"
                    for gender, count in gender_counts.items()
                    if count > 0
                )
                or "None"
            )

            average_age = (
                statistics.mean(
                    _age_on_date(leader.birthday, day.date) for leader in responsibles
                )
                if responsibles
                else 0.0
            )

            entry_metrics.append(
                {
                    "day": day,
                    "entry": entry,
                    "responsible_count": responsible_count,
                    "first_year_count": first_year_count,
                    "returning_leader_count": returning_leader_count,
                    "management_count": management_count,
                    "average_age": average_age,
                    "gender_mix": gender_mix,
                    "has_mixed_gender": sum(
                        count > 0 for count in gender_counts.values()
                    )
                    > 1,
                    "has_experienced_cover": (returning_leader_count + management_count)
                    > 0,
                }
            )

    return entry_metrics


def fill_metrics_sheet(camp: Camp, ws: Worksheet):
    _setup_worksheet(ws)

    title = ws.cell(row=1, column=1, value="Schedule Quality Metrics")
    title.font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    leader_load = _calculate_leader_load(camp)
    loads = [metrics["total"] for metrics in leader_load]
    max_tasks_on_day = max(
        (metrics["max_tasks_on_day"] for metrics in leader_load), default=0
    )
    max_consecutive_active = max(
        (metrics["max_consecutive_active_days"] for metrics in leader_load),
        default=0,
    )
    leaders_with_no_day_off = sum(
        1 for metrics in leader_load if metrics["day_offs"] == 0
    )
    leaders_with_long_streak = sum(
        1 for metrics in leader_load if metrics["max_consecutive_active_days"] >= 4
    )

    spacing = _calculate_responsibility_spacing(camp)
    spacing_by_leader = {
        metrics["leader_name"]: metrics for metrics in spacing["leader_spacing"]
    }
    all_spacing_gaps = spacing["all_gap_minutes"]
    avg_spacing_gap_hours = (
        (statistics.mean(all_spacing_gaps) / 60) if all_spacing_gaps else 0.0
    )
    leaders_with_dense_spacing = sum(
        1
        for metrics in spacing["leader_spacing"]
        if metrics["gap_count"] > 0 and metrics["tight_gap_ratio"] >= 0.5
    )
    spacing_type_labels = ", ".join(
        sorted(entry_type.value.short for entry_type in spacing["included_entry_types"])
    )

    eligibility = _calculate_eligibility(camp)
    eligible_counts = [item["eligible_count"] for item in eligibility]
    total_entries = len(eligibility)
    entries_with_two_or_less = sum(
        1 for item in eligibility if item["eligible_count"] <= 2
    )
    entries_with_zero = sum(1 for item in eligibility if item["eligible_count"] == 0)
    average_eligible = statistics.mean(eligible_counts) if eligible_counts else 0.0
    min_eligible = min(eligible_counts, default=0)
    max_eligible = max(eligible_counts, default=0)

    interactions = _calculate_interactions(camp)
    unique_partner_counts = list(interactions["unique_partner_counts"].values())
    pair_counts = list(interactions["pair_counts"].values())
    avg_unique_partners = (
        statistics.mean(unique_partner_counts) if unique_partner_counts else 0.0
    )
    max_pair_count = max(pair_counts, default=0)
    pair_stddev = _safe_stdev(pair_counts)

    entry_assignment = _calculate_entry_assignment_metrics(camp)
    avg_responsibles_per_entry = (
        statistics.mean(item["responsible_count"] for item in entry_assignment)
        if entry_assignment
        else 0.0
    )
    entries_with_management = sum(
        1 for item in entry_assignment if item["management_count"] > 0
    )
    entries_without_experienced_cover = sum(
        1
        for item in entry_assignment
        if item["responsible_count"] > 0 and not item["has_experienced_cover"]
    )
    entries_with_mixed_gender = sum(
        1 for item in entry_assignment if item["has_mixed_gender"]
    )
    avg_entry_responsible_age = (
        statistics.mean(
            item["average_age"]
            for item in entry_assignment
            if item["responsible_count"] > 0
        )
        if any(item["responsible_count"] > 0 for item in entry_assignment)
        else 0.0
    )

    summary_rows = [
        ("Total leaders", len(camp.leaders_incl_management)),
        ("Responsible entries", total_entries),
        ("Total assigned responsibilities", sum(loads)),
        (
            "Average responsibilities per leader",
            round(statistics.mean(loads), 2) if loads else 0.0,
        ),
        (
            "Average responsibles per required entry",
            round(avg_responsibles_per_entry, 2),
        ),
        ("Responsibility stdev", round(_safe_stdev(loads), 2)),
        ("Maximum tasks on one day", max_tasks_on_day),
        ("Entries with <=2 eligible leaders", entries_with_two_or_less),
        ("Entries with 0 eligible leaders", entries_with_zero),
        ("Average eligible leaders per entry", round(average_eligible, 2)),
        ("Min eligible leaders", min_eligible),
        ("Max eligible leaders", max_eligible),
        ("Entries with camp management assigned", entries_with_management),
        (
            "Entries without experienced cover",
            entries_without_experienced_cover,
        ),
        ("Entries with mixed-gender responsibles", entries_with_mixed_gender),
        (
            "Average responsible age per entry",
            round(avg_entry_responsible_age, 2),
        ),
        ("Average unique partners per leader", round(avg_unique_partners, 2)),
        ("Max shared tasks for a pair", max_pair_count),
        ("Pair count stdev", round(pair_stddev, 2)),
        ("Leaders with 0 day offs", leaders_with_no_day_off),
        ("Max consecutive active days", max_consecutive_active),
        ("Leaders with 4+ consecutive active days", leaders_with_long_streak),
        (
            "Average gap between selected responsibilities (h)",
            round(avg_spacing_gap_hours, 2),
        ),
        (
            "Leaders with >=50% tight gaps (<=2h)",
            leaders_with_dense_spacing,
        ),
        ("Spacing entry types", spacing_type_labels),
    ]

    row = 3
    for label, value in summary_rows:
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    # Risk table for constrained tasks
    risk_title = ws.cell(row=3, column=5, value="Constrained entries")
    risk_title.font = Font(bold=True)
    risk_start_row = 4
    headers = ["Day", "Entry", "Responsible", "Eligible leaders"]
    _write_table_headers(ws, risk_start_row, 5, headers)

    risk_rows = sorted(
        eligibility,
        key=lambda item: (item["eligible_count"], item["responsible_count"]),
    )[:15]
    for offset, item in enumerate(risk_rows, start=1):
        row_index = risk_start_row + offset
        ws.cell(row=row_index, column=5, value=item["day"].as_str())
        ws.cell(row=row_index, column=6, value=item["entry"].entry_type.value.short)
        ws.cell(row=row_index, column=7, value=item["responsible_count"])
        eligible_cell = ws.cell(row=row_index, column=8, value=item["eligible_count"])
        if item["eligible_count"] == 0:
            eligible_cell.fill = PatternFill(
                start_color=Colors.LIGHT_PURPLE,
                end_color=Colors.LIGHT_PURPLE,
                fill_type="solid",
            )
        elif item["eligible_count"] <= 2:
            eligible_cell.fill = PatternFill(
                start_color=Colors.LIGHT_YELLOW,
                end_color=Colors.LIGHT_YELLOW,
                fill_type="solid",
            )

    # Entry allocation profile table
    entry_title = ws.cell(row=3, column=10, value="Entry allocation profile")
    entry_title.font = Font(bold=True)
    entry_start_row = 4
    entry_headers = [
        "Day",
        "Entry",
        "Responsible",
        "First-year",
        "Returning",
        "Mgmt",
        "Avg age",
        "Gender mix",
    ]
    _write_table_headers(ws, entry_start_row, 10, entry_headers)

    sorted_entry_metrics = sorted(
        entry_assignment,
        key=lambda item: (
            item["day"].date,
            item["entry"].start_time or "",
            item["entry"].entry_type.value.short,
        ),
    )
    for offset, item in enumerate(sorted_entry_metrics, start=1):
        row_index = entry_start_row + offset
        ws.cell(row=row_index, column=10, value=item["day"].as_str())
        ws.cell(row=row_index, column=11, value=item["entry"].entry_type.value.short)
        responsible_cell = ws.cell(
            row=row_index, column=12, value=item["responsible_count"]
        )
        ws.cell(row=row_index, column=13, value=item["first_year_count"])
        ws.cell(row=row_index, column=14, value=item["returning_leader_count"])
        ws.cell(row=row_index, column=15, value=item["management_count"])
        ws.cell(row=row_index, column=16, value=round(item["average_age"], 1))
        ws.cell(row=row_index, column=17, value=item["gender_mix"])

        if item["responsible_count"] == 0:
            responsible_cell.fill = PatternFill(
                start_color=Colors.LIGHT_PURPLE,
                end_color=Colors.LIGHT_PURPLE,
                fill_type="solid",
            )
        elif not item["has_experienced_cover"]:
            responsible_cell.fill = PatternFill(
                start_color=Colors.LIGHT_YELLOW,
                end_color=Colors.LIGHT_YELLOW,
                fill_type="solid",
            )

    # Leader workload table
    load_title_row = max(risk_start_row + len(risk_rows) + 2, row + 1, 20)
    ws.cell(row=load_title_row, column=1, value="Leader workload")
    ws.cell(row=load_title_row, column=1).font = Font(bold=True)

    workload_headers = [
        "Leader",
        "Total",
        "Active days",
        "Max/day",
        "Consecutive days",
        "Day offs",
        "Avg gap (h)",
        "Tight gaps <=2h (%)",
    ]
    _write_table_headers(ws, load_title_row + 1, 1, workload_headers)

    for offset, metrics in enumerate(
        sorted(leader_load, key=lambda item: item["leader_name"]), start=1
    ):
        row_index = load_title_row + 1 + offset
        ws.cell(row=row_index, column=1, value=metrics["leader_name"])
        ws.cell(row=row_index, column=2, value=metrics["total"])
        ws.cell(row=row_index, column=3, value=metrics["active_days"])
        ws.cell(row=row_index, column=4, value=metrics["max_tasks_on_day"])
        ws.cell(row=row_index, column=5, value=metrics["max_consecutive_active_days"])
        ws.cell(row=row_index, column=6, value=metrics["day_offs"])
        spacing_metrics = spacing_by_leader.get(metrics["leader_name"])
        avg_gap_hours = (
            round(spacing_metrics["avg_gap_hours"], 2) if spacing_metrics else 0.0
        )
        tight_gap_percent = (
            round(spacing_metrics["tight_gap_ratio"] * 100, 1)
            if spacing_metrics
            else 0.0
        )
        ws.cell(row=row_index, column=7, value=avg_gap_hours)
        ws.cell(row=row_index, column=8, value=tight_gap_percent)

    _auto_adjust_column_widths(ws)
