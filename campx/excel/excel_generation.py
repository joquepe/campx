from campx.model.camp import Camp
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties, WorksheetProperties
from campx.excel.responsibilities import fill_responsibilities_sheet
from campx.excel.schedule import (
    fill_schedule_sheet,
    fill_schedule_sheet_for_participant,
    make_unique_schedule_sheet_name,
)
from campx.excel.overview import fill_overview_sheet
from campx.excel.eligible_leaders import fill_eligible_leaders_sheet
from campx.excel.leader_interactions import fill_leader_interactions_sheet
from campx.excel.metrics import fill_metrics_sheet
from campx.excel.working_hours import fill_working_hours_sheet
from campx.validation import get_errors
from openpyxl.worksheet.worksheet import Worksheet

import logging

logger = logging.getLogger(__name__)


def add_schedule_sheet(camp: Camp, workbook: Workbook):
    logger.info("Adding schedule sheet...")
    schedule_sheet = workbook.create_sheet(title="Schema")
    fill_schedule_sheet(camp, schedule_sheet)


def add_responsibilities_sheet(camp: Camp, workbook: Workbook):
    logger.info("Adding responsibilities sheet...")
    responsibilities_sheet = workbook.create_sheet(title="Ansvar")
    fill_responsibilities_sheet(camp, responsibilities_sheet)


def add_overview_sheet(camp: Camp, workbook: Workbook):
    logger.info("Adding overview sheet...")
    overview_sheet = workbook.create_sheet(title="Överblick")
    fill_overview_sheet(camp, overview_sheet)


def add_metrics_sheet(camp: Camp, workbook: Workbook):
    logger.info("Adding metrics sheet...")
    metrics_sheet = workbook.create_sheet(title="Kvalitet")
    fill_metrics_sheet(camp, metrics_sheet)


def add_eligible_leaders_by_schedule_entry(camp: Camp, workbook: Workbook):
    logger.info("Adding eligible leaders sheet...")
    eligible_leaders_sheet = workbook.create_sheet(title="Möjliga")
    fill_eligible_leaders_sheet(camp, eligible_leaders_sheet)


def add_leader_interactions_sheet(camp: Camp, workbook: Workbook):
    logger.info("Adding leader interactions sheet...")
    interactions_sheet = workbook.create_sheet(title="Interaktioner")
    fill_leader_interactions_sheet(camp, interactions_sheet)


def add_working_hours_sheet(camp: Camp, workbook: Workbook):
    logger.info("Adding working hours sheet...")
    working_hours_sheet = workbook.create_sheet(title="Arbetstimmar")
    fill_working_hours_sheet(camp, working_hours_sheet)


def add_validation_errors_sheet(camp: Camp, workbook: Workbook):
    logger.info("Adding validation errors sheet...")
    validation_sheet = workbook.create_sheet(title="Valideringsfel")
    fill_validation_errors_sheet(camp, validation_sheet)


def fill_validation_errors_sheet(camp: Camp, ws: Worksheet):
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
    ws.print_options.gridLines = True
    ws.print_options.horizontalCentered = True
    ws.sheet_properties = WorksheetProperties(
        pageSetUpPr=PageSetupProperties(fitToPage=True)
    )

    ws.cell(row=1, column=1, value="Validation errors")
    ws.cell(row=1, column=1).font = Font(bold=True)
    severity_header = ws.cell(row=2, column=1, value="Severity")
    severity_header.font = Font(bold=True)
    error_header = ws.cell(row=2, column=2, value="Error")
    error_header.font = Font(bold=True)

    errors = get_errors(camp)
    if not errors:
        ws.cell(row=3, column=2, value="No validation errors found.")
    else:
        for row, error in enumerate(errors, start=3):
            ws.cell(row=row, column=1, value=error.severity)
            ws.cell(row=row, column=2, value=str(error))

    for column in (1, 2):
        max_len = max(
            len(str(ws.cell(row=r, column=column).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[get_column_letter(column)].width = max(20, max_len + 2)


def add_participant_schedule_sheets(camp: Camp, workbook: Workbook):
    logger.info("Adding participant schedule sheets...")
    for participant in camp.participants:
        sheet_name = make_unique_schedule_sheet_name(workbook, participant)
        participant_sheet = workbook.create_sheet(title=sheet_name)
        fill_schedule_sheet_for_participant(camp, participant_sheet, participant)


def create_camp_excel(camp: Camp):
    logger.info("Generating excel...")
    workbook = Workbook()
    sheets = [
        add_schedule_sheet,
        add_responsibilities_sheet,
        add_working_hours_sheet,
        add_overview_sheet,
        add_metrics_sheet,
        add_eligible_leaders_by_schedule_entry,
        add_leader_interactions_sheet,
        add_validation_errors_sheet,
        add_participant_schedule_sheets,
    ]
    for sheet in sheets:
        sheet(camp, workbook)

    # Save the workbook
    output_path = Path()
    workbook.save(f"{output_path}/{camp.name}_schema.xlsx")
    logger.info(f"Excel saved at {output_path}/{camp.name}_schema.xlsx")
