from campx.model.camp import Camp
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from campx.model.enums import ParticipantType, EntryType
from campx.model.schedule_entry import ScheduleEntry
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from campx.excel.utilities import get_activity_color
import datetime as dt
from campx.excel.utilities import Colors


def sort_by_start_time(entry: ScheduleEntry):
    if entry.entry_type == EntryType.NIGHT_DUTY:
        return (1, dt.datetime.strptime("23:59", "%H:%M").time())
    if entry.start_time is None:
        return (0, None)
    time_obj = dt.datetime.strptime(entry.start_time, "%H:%M").time()
    return (1, time_obj)


def fill_responsibilities_sheet(camp: Camp, ws: Worksheet):
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToPage = True

    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

    ws.print_options.gridLines = True
    ws.print_options.horizontalCentered = True

    ws.sheet_properties = WorksheetProperties(
        pageSetUpPr=PageSetupProperties(fitToPage=True)
    )

    day_col_widths = {}
    for day in camp.schedule.days:
        max_tasks = max(
            len(leader.get_responsible_entries(camp.schedule, day))
            for leader in camp.leaders_incl_management
        )
        day_col_widths[day.as_str()] = max_tasks or 1  # At least 1 column per day

    day_start_cols = {}
    col = 2
    for day in camp.schedule.days:
        day_start_cols[day.as_str()] = col
        col += day_col_widths[day.as_str()]

    ws.cell(row=1, column=1, value="Namn")
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=1).border = Border(
        bottom=Side(style="medium", color=Colors.BLACK),
        right=Side(style="thin", color=Colors.BLACK),
    )
    ws.cell(row=1, column=1).alignment = Alignment(
        vertical="center", horizontal="center"
    )
    for day in camp.schedule.days:
        start_col = day_start_cols[day.as_str()]
        span = day_col_widths[day.as_str()]
        end_col = start_col + span - 1
        ws.merge_cells(
            start_row=1, start_column=start_col, end_row=1, end_column=end_col
        )
        ws.cell(
            row=1,
            column=start_col,
            value=day.as_str(
                format="%a %d %b",
                swedish_month_names=True,
                swedish_weekday_names=True,
            ),
        )
        ws.cell(row=1, column=start_col).alignment = Alignment(
            vertical="center", horizontal="center"
        )
        ws.cell(row=1, column=start_col).font = Font(bold=True)
        for col in range(start_col, end_col + 1):
            if col == end_col:
                border = Border(
                    bottom=Side(style="medium", color=Colors.BLACK),
                    right=Side(style="thin", color=Colors.BLACK),
                )
            else:
                border = Border(bottom=Side(style="medium", color=Colors.BLACK))
            ws.cell(row=1, column=col).border = border

    row = 2
    for leader in camp.leaders_incl_management:
        ws.cell(row=row, column=1, value=leader.full_name)
        if leader.participant_type == ParticipantType.CAMP_MANAGEMENT:
            ws.cell(row=row, column=1).font = Font(
                bold=True, italic=True, color=Colors.GREY
            )
        else:
            ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).border = Border(
            bottom=Side(style="thin", color=Colors.GREY),
            right=Side(style="thin", color=Colors.BLACK),
        )

        for day in camp.schedule.days:
            entries = leader.get_responsible_entries(camp.schedule, day)
            entries = sorted(entries, key=sort_by_start_time)
            tasks = {
                e.entry_type.value.short: get_activity_color(e.entry_type)
                for e in entries
            }
            start_col = day_start_cols[day.as_str()]
            for i, (task, color) in enumerate(tasks.items()):
                cell = ws.cell(row=row, column=start_col + i, value=task)
                cell.fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
                cell.font = Font(italic=True)
                cell.alignment = Alignment(vertical="center", horizontal="center")
            span = day_col_widths[day.as_str()]
            end_col = start_col + span - 1
            for col in range(start_col, end_col + 1):
                if col == end_col:
                    border = Border(
                        bottom=Side(style="thin", color=Colors.GREY),
                        right=Side(style="thin", color=Colors.BLACK),
                    )
                else:
                    border = Border(
                        bottom=Side(style="thin", color=Colors.GREY),
                        right=Side(style="thin", color=Colors.BLACK),
                    )

                ws.cell(row=row, column=col).border = border

        row += 1

    for col in range(1, ws.max_column + 1):
        max_len = max(
            (
                len(str(ws.cell(row=r, column=col).value or ""))
                for r in range(1, ws.max_row + 1)
            ),
            default=0,
        )
        ws.column_dimensions[get_column_letter(col)].width = max(10, max_len + 2)
