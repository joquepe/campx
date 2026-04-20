from campx.model.camp import Camp
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from campx.validation import is_eligible_participant
from campx.excel.utilities import get_activity_color
from openpyxl.comments import Comment
from campx.excel.utilities import Colors


def fill_eligible_leaders_sheet(camp: Camp, ws: Worksheet):
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = "landscape"  # Optional
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 0  # Use 1 if you also want to fit to 1 page wide

    # Optional: Set margins and scaling if needed
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

    # Optional: Print gridlines if you want the cell lines visible
    ws.print_options.gridLines = True
    ws.print_options.horizontalCentered = True

    # Optional: Make sure worksheet is printable
    ws.sheet_properties = WorksheetProperties(
        pageSetUpPr=PageSetupProperties(fitToPage=True)
    )

    ws.cell(row=1, column=1, value="Day")
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=1).border = Border(
        bottom=Side(style="medium", color=Colors.BLACK),
        right=Side(style="thin", color=Colors.BLACK),
    )

    ws.cell(row=1, column=2, value="Entry Type")
    ws.cell(row=1, column=2).font = Font(bold=True)
    ws.cell(row=1, column=2).border = Border(
        bottom=Side(style="medium", color=Colors.BLACK),
        right=Side(style="thin", color=Colors.BLACK),
    )

    col = 3
    for leader in camp.leaders_incl_management:
        ws.cell(row=1, column=col, value=leader.initials)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).border = Border(
            bottom=Side(style="medium", color=Colors.BLACK),
            right=Side(style="thin", color=Colors.BLACK),
        )
        col += 1

    row = 2
    for day in camp.schedule.days:
        for entry in day.schedule_entries:
            ws.cell(row=row, column=1, value=day.as_str())
            ws.cell(row=row, column=2, value=entry.entry_type.value.short)
            entry_color = get_activity_color(entry.entry_type)
            ws.cell(row=row, column=2).fill = PatternFill(
                start_color=entry_color, end_color=entry_color, fill_type="solid"
            )
            ws.cell(row=row, column=1).border = Border(
                bottom=Side(style="thin", color=Colors.BLACK),
                right=Side(style="thin", color=Colors.BLACK),
            )
            ws.cell(row=row, column=2).border = Border(
                bottom=Side(style="thin", color=Colors.BLACK),
                right=Side(style="thin", color=Colors.BLACK),
            )

            for col, leader in enumerate(camp.leaders_incl_management, start=3):
                eligible = False
                comment = None
                if leader in entry.responsible:
                    color = Colors.BLUE
                else:
                    eligible, comment = is_eligible_participant(
                        leader, entry, camp, day
                    )
                    color = "39FF14" if eligible else "FF073A"

                if not eligible and comment:
                    ws.cell(row=row, column=col).comment = Comment(
                        text=comment, author=""
                    )

                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
                ws.cell(row=row, column=col).border = Border(
                    bottom=Side(style="thin", color=Colors.BLACK),
                    right=Side(style="thin", color=Colors.BLACK),
                )

            row += 1
        # add border along entire row after each day
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row - 1, column=col).border = Border(
                bottom=Side(style="medium", color=Colors.BLACK),
                right=Side(style="thin", color=Colors.BLACK),
            )
