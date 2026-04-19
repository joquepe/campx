from campx.model.camp import Camp
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from collections import defaultdict
from itertools import combinations
from campx.model.enums import EntryType
from campx.excel.utilities import Colors

EXCLUDE_ENTRY_TYPES = [EntryType.BOAT, EntryType.EDUCATION, EntryType.FEEDBACK]


def fill_leader_interactions_sheet(camp: Camp, ws: Worksheet):
    """
    Fill a sheet showing how many times each pair of leaders works together.
    """
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1

    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
    ws.print_options.gridLines = True
    ws.print_options.horizontalCentered = True

    ws.sheet_properties = WorksheetProperties(
        pageSetUpPr=PageSetupProperties(fitToPage=True)
    )

    leaders = camp.leaders_incl_management
    leader_list = sorted(leaders, key=lambda p: p.full_name)

    # Build a matrix to store counts and entry type distributions
    # Key: (leader1_idx, leader2_idx), Value: dict with entry types and counts
    interaction_matrix = defaultdict(lambda: defaultdict(int))

    # Iterate through all schedule entries
    for day in camp.schedule.days:
        for entry in day.schedule_entries:
            if entry.entry_type in EXCLUDE_ENTRY_TYPES:
                continue
            responsible = entry.responsible
            # For each unique unordered pair of leaders in this entry
            for leader1, leader2 in combinations(responsible, 2):
                idx1 = leader_list.index(leader1)
                idx2 = leader_list.index(leader2)
                key = (min(idx1, idx2), max(idx1, idx2))
                interaction_matrix[key][entry.entry_type.value.short] += 1

    # Set up header row with leader initials (columns)
    ws.cell(row=1, column=1, value="Leaders")
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=1).border = Border(
        bottom=Side(style="medium", color=Colors.BLACK),
        right=Side(style="medium", color=Colors.BLACK),
    )
    ws.cell(row=1, column=1).alignment = Alignment(
        horizontal="center", vertical="center"
    )

    for col, leader in enumerate(leader_list, start=2):
        ws.cell(row=1, column=col, value=leader.initials)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).border = Border(
            bottom=Side(style="medium", color=Colors.BLACK),
            right=Side(style="thin", color=Colors.BLACK),
        )
        ws.cell(row=1, column=col).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    # Set up header column with leader initials (rows)
    for row, leader in enumerate(leader_list, start=2):
        ws.cell(row=row, column=1, value=leader.initials)
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).border = Border(
            bottom=Side(style="thin", color=Colors.BLACK),
            right=Side(style="medium", color=Colors.BLACK),
        )
        ws.cell(row=row, column=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )

    # Fill in the interaction counts
    for row, leader1 in enumerate(leader_list, start=2):
        for col, leader2 in enumerate(leader_list, start=2):
            idx1 = leader_list.index(leader1)
            idx2 = leader_list.index(leader2)

            if idx1 == idx2:
                # Diagonal: count self entries (leader paired with self)
                count = 0
                self_entries = defaultdict(int)
                for day in camp.schedule.days:
                    for entry in day.schedule_entries:
                        if leader1 in entry.responsible:
                            count += 1
                            self_entries[entry.entry_type.value.short] += 1

                cell = ws.cell(row=row, column=col, value=count)
                if self_entries:
                    comment_text = ""
                    for entry_type in sorted(self_entries.keys()):
                        comment_text += f"{entry_type}: {self_entries[entry_type]}\n"
                    cell.comment = Comment(text=comment_text.strip(), author="")
            else:
                # Off-diagonal: count pairs
                key_forward = (min(idx1, idx2), max(idx1, idx2))
                count = sum(interaction_matrix[key_forward].values())
                cell = ws.cell(row=row, column=col, value=count if count > 0 else "")

                # Create comment with entry type distribution
                if count > 0:
                    comment_text = ""
                    for entry_type in sorted(interaction_matrix[key_forward].keys()):
                        comment_text += f"{entry_type}: {interaction_matrix[key_forward][entry_type]}\n"
                    cell.comment = Comment(text=comment_text.strip(), author="")

            # Style the cell
            if count > 0 and (idx1 != idx2):
                # Color based on count intensity
                color = _get_intensity_color(count)
                cell.fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
            else:
                cell.fill = PatternFill(
                    start_color=Colors.LIGHT_LIGHT_GREY,
                    end_color=Colors.LIGHT_LIGHT_GREY,
                    fill_type="solid",
                )

            cell.border = Border(
                bottom=Side(style="thin", color=Colors.BLACK),
                right=Side(style="thin", color=Colors.BLACK),
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set column widths
    ws.column_dimensions["A"].width = 12
    for col_idx in range(2, len(leader_list) + 2):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 10


def _get_intensity_color(count: int) -> str:
    """Get a color based on the count intensity (lighter = lower count, darker = higher count)."""
    if count <= 1:
        return "E8F5E9"  # Very light green
    elif count <= 3:
        return "C8E6C9"  # Light green
    elif count <= 5:
        return "A1887F"  # Brownish
    elif count <= 8:
        return "FFB74D"  # Orange
    elif count <= 12:
        return "FF5722"  # Red-orange
    else:
        return "F44336"  # Dark red
