"""Participant-specific schedule sheets with highlighting."""

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

from campx.model.camp import Camp
from campx.model.participant import Participant
from campx.model.enums import EntryType
from campx.excel.schedule import (
    fill_schedule_sheet,
    get_time_slots,
    find_cell_for_activity,
    _make_unique_sheet_name,
    _get_secondary_row,
    START_HOUR,
    END_HOUR,
)


def fill_schedule_sheet_for_participant(
    camp: Camp, worksheet: Worksheet, participant: Participant
):
    """Create a participant-specific schedule sheet with highlighting."""
    fill_schedule_sheet(camp, worksheet)
    participant_label = participant.nick_name or participant.full_name
    label_col = worksheet.max_column + 2
    label_cell = worksheet.cell(
        row=1, column=label_col, value=f"Participant: {participant_label}"
    )
    label_cell.font = Font(bold=True, color="FF0000")
    _highlight_responsible_cells_for_participant_sheet(camp, worksheet, participant)


def _highlight_responsible_cells_for_participant_sheet(
    camp: Camp, worksheet: Worksheet, participant: Participant
):
    """Highlight responsible-name cells only where the sheet participant is responsible."""
    participant_token = (
        participant.nick_name or participant.initials or participant.full_name
    )
    top_types = EntryType.time_independent_top()
    bottom_types = EntryType.time_independent_bottom()
    time_slots = get_time_slots(START_HOUR, END_HOUR, len(top_types))
    bottom_start_row = max(time_slots.values()) + 2

    for col, day in enumerate(camp.schedule.days, start=2):
        for row_offset, entry_type in enumerate(top_types, start=1):
            entries = day.get_entries(entry_type)
            if entries and any(participant in entry.responsible for entry in entries):
                cell = worksheet.cell(row=row_offset, column=col)
                _highlight_token_in_cell(cell, participant_token)

        timed_entries = [
            entry
            for entry in day.schedule_entries
            if not entry.entry_type.value.time_independent
            and entry.entry_type.value.show_in_schedule
            and participant in entry.responsible
        ]
        for activity in timed_entries:
            cell, _ = find_cell_for_activity(time_slots, activity, worksheet, col)
            secondary_cell = worksheet.cell(
                row=_get_secondary_row(cell.row), column=col
            )
            _highlight_token_in_cell(secondary_cell, participant_token)

        for row_offset, entry_type in enumerate(bottom_types, start=1):
            entries = day.get_entries(entry_type)
            if entries and any(participant in entry.responsible for entry in entries):
                cell = worksheet.cell(row=bottom_start_row + row_offset - 1, column=col)
                _highlight_token_in_cell(cell, participant_token)


def _highlight_token_in_cell(cell, token: str) -> None:
    """Apply red bold styling only to exact token occurrences inside a cell value."""
    if not token:
        return
    value = cell.value
    if not isinstance(value, str) or token not in value:
        return

    pieces = value.split(token)
    rich_text = CellRichText()
    highlight_font = InlineFont(b=True, color="00FF0000")
    for idx, piece in enumerate(pieces):
        if piece:
            rich_text.append(piece)
        if idx < len(pieces) - 1:
            rich_text.append(TextBlock(highlight_font, token))
    cell.value = rich_text


def make_unique_schedule_sheet_name(workbook: Workbook, participant: Participant):
    """Generate a unique sheet name for the participant schedule sheet."""
    return _make_unique_sheet_name(
        workbook, participant.nick_name or participant.full_name
    )
