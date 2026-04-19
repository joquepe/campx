from campx.excel.schedule import find_cell_for_activity
from campx.excel.schedule import fill_schedule_sheet
from campx.excel.responsibilities import fill_responsibilities_sheet
from campx.model.schedule_entry import ScheduleEntry
from campx.model.enums import EntryType
from campx.model.enums import ParticipantType
from campx.model.participant import Participant
from campx.model.camp import Camp
from campx.model.camp_place import CampPlace
from campx.model.day import Day
from campx.model.schedule import Schedule
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class TestExcelScheduleGeneration:
    """Test suite for Excel schedule generation functionality."""

    def test_responsibilities_sheet_uses_landscape_a4_page_setup(self):
        wb = Workbook()
        ws = wb.active

        leader = Participant(
            participant_id=1,
            first_name="Alice",
            last_name="Andersson",
            gender="F",
            birthday="1990-01-15",
            participant_type=ParticipantType.LEADER,
        )

        day = Day(
            date=__import__("datetime").date(2026, 4, 12),
            schedule_entries=[
                ScheduleEntry(
                    entry_type=EntryType.MORNING_PRAYER,
                    name="",
                    start_time="08:00",
                    end_time="08:30",
                    responsible=[leader],
                )
            ],
        )
        camp = Camp(
            name="TestCamp",
            camp_place=CampPlace("TestPlace"),
            participants=[leader],
            schedule=Schedule(days=[day]),
        )

        fill_responsibilities_sheet(camp, ws)

        assert str(ws.page_setup.paperSize) == "9"  # A4
        assert ws.page_setup.orientation == "landscape"
        assert ws.page_setup.fitToWidth == 1
        assert ws.page_setup.fitToHeight == 1
        assert ws.sheet_properties.pageSetUpPr.fitToPage is True

    def test_time_dependent_activity_uses_custom_name_when_available(self):
        """Test that time-dependent activities use custom name when available."""
        # Create a mock worksheet
        wb = Workbook()
        ws = wb.active

        # Create a time-dependent activity with custom name
        entry = ScheduleEntry(
            entry_type=EntryType.MORNING_PRAYER,
            name="Custom Morning Prayer Name",
            start_time="08:00",
            end_time="08:30",
        )

        # Mock time slots
        time_slots = {"08:00": 5, "08:30": 6}

        # Call the function
        cell, value_to_cell = find_cell_for_activity(time_slots, entry, ws, 2)

        # Should use the custom name
        assert "Custom Morning Prayer Name" in value_to_cell
        assert EntryType.MORNING_PRAYER.value.long not in value_to_cell

    def test_time_dependent_activity_uses_long_name_when_no_custom_name(self):
        """Test that time-dependent activities use long name when no custom name."""
        wb = Workbook()
        ws = wb.active

        # Create a time-dependent activity without custom name
        entry = ScheduleEntry(
            entry_type=EntryType.LUNCH,
            name="",  # No custom name
            start_time="12:00",
            end_time="13:00",
        )

        time_slots = {"12:00": 10}

        cell, value_to_cell = find_cell_for_activity(time_slots, entry, ws, 2)

        # Should use the long name
        assert EntryType.LUNCH.value.long in value_to_cell

    def test_time_dependent_activity_with_responsible(self):
        """Test time-dependent activity with responsible participants."""
        wb = Workbook()
        ws = wb.active

        # Create participants
        p1 = Participant(
            participant_id=1,
            first_name="John",
            last_name="Doe",
            gender="M",
            birthday="1990-01-15",
            participant_type="LEADER",
        )

        entry = ScheduleEntry(
            entry_type=EntryType.LUNCH,
            name="Special Lunch",
            start_time="12:00",
            end_time="13:00",
            responsible=[p1],
        )

        time_slots = {"12:00": 10}

        cell, value_to_cell = find_cell_for_activity(time_slots, entry, ws, 2)

        # Should use custom name and include responsible initials
        assert "Special Lunch" in value_to_cell
        assert "JD" in value_to_cell  # John Doe's initials

    def test_time_dependent_activity_with_additional_info(self):
        """Test time-dependent activity appends additional info after responsibles."""
        wb = Workbook()
        ws = wb.active

        p1 = Participant(
            participant_id=1,
            first_name="John",
            last_name="Doe",
            gender="M",
            birthday="1990-01-15",
            participant_type="LEADER",
        )

        entry = ScheduleEntry(
            entry_type=EntryType.LUNCH,
            name="Special Lunch",
            start_time="12:00",
            end_time="13:00",
            responsible=[p1],
            additional_info="Bring napkins",
        )

        time_slots = {"12:00": 10}

        cell, value_to_cell = find_cell_for_activity(time_slots, entry, ws, 2)

        assert "Special Lunch" in value_to_cell
        assert "JD" in value_to_cell
        assert "Bring napkins" in value_to_cell

    def test_time_dependent_activity_with_start_time_display(self):
        """Test that start time is shown when it doesn't match the slot."""
        wb = Workbook()
        ws = wb.active

        entry = ScheduleEntry(
            entry_type=EntryType.EVENING_ACTIVITY,
            name="Custom Activity",
            start_time="19:15",  # Doesn't match slot time
            end_time="20:00",
        )

        time_slots = {"19:00": 15}

        cell, value_to_cell = find_cell_for_activity(time_slots, entry, ws, 2)

        # Should include the actual start time
        assert "Custom Activity" in value_to_cell
        assert "19:15" in value_to_cell

    def test_schedule_day_columns_and_hour_rows_are_uniform(self):
        wb = Workbook()
        ws = wb.active

        leader = Participant(
            participant_id=1,
            first_name="Alice",
            last_name="Andersson",
            gender="F",
            birthday="1990-01-15",
            participant_type=ParticipantType.LEADER,
        )

        day1 = Day(
            date=__import__("datetime").date(2026, 4, 12),
            schedule_entries=[
                ScheduleEntry(
                    entry_type=EntryType.MORNING_PRAYER,
                    name="",
                    start_time="08:00",
                    end_time="08:30",
                    responsible=[leader],
                )
            ],
        )
        day2 = Day(
            date=__import__("datetime").date(2026, 4, 13),
            schedule_entries=[
                ScheduleEntry(
                    entry_type=EntryType.EVENING_ACTIVITY,
                    name="",
                    start_time="19:00",
                    end_time="20:00",
                    responsible=[leader],
                )
            ],
        )
        camp = Camp(
            name="TestCamp",
            camp_place=CampPlace("TestPlace"),
            participants=[leader],
            schedule=Schedule(days=[day1, day2]),
        )

        fill_schedule_sheet(camp, ws)

        day_col_widths = [
            ws.column_dimensions[get_column_letter(col)].width for col in range(2, 4)
        ]
        assert len(set(day_col_widths)) == 1

        hour_rows = []
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=1).value
            if isinstance(value, str) and ":" in value:
                hour_rows.append(row)

        hour_heights = [ws.row_dimensions[row].height for row in hour_rows]
        assert hour_heights
        assert len(set(hour_heights)) == 1

    def test_responsible_names_are_italic_in_schedule(self):
        wb = Workbook()
        ws = wb.active

        leader = Participant(
            participant_id=1,
            first_name="Alice",
            last_name="Andersson",
            gender="F",
            birthday="1990-01-15",
            participant_type=ParticipantType.LEADER,
            nick_name="Ali",
        )

        day = Day(
            date=__import__("datetime").date(2026, 4, 12),
            schedule_entries=[
                ScheduleEntry(
                    entry_type=EntryType.MORNING_PRAYER,
                    name="Morning Prayer",
                    start_time="08:00",
                    end_time="08:30",
                    responsible=[leader],
                )
            ],
        )
        camp = Camp(
            name="TestCamp",
            camp_place=CampPlace("TestPlace"),
            participants=[leader],
            schedule=Schedule(days=[day]),
        )

        fill_schedule_sheet(camp, ws)

        name_cell = None
        responsible_cell = None
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "Morning Prayer" in cell.value:
                    name_cell = cell
                if isinstance(cell.value, str) and "Ali" in cell.value:
                    responsible_cell = cell

        assert name_cell is not None
        assert responsible_cell is not None
        assert responsible_cell.row == name_cell.row + 1
        assert responsible_cell.font and responsible_cell.font.italic

    def test_named_entries_are_bold_in_schedule(self):
        wb = Workbook()
        ws = wb.active

        leader = Participant(
            participant_id=1,
            first_name="Alice",
            last_name="Andersson",
            gender="F",
            birthday="1990-01-15",
            participant_type=ParticipantType.LEADER,
            nick_name="Ali",
        )

        day = Day(
            date=__import__("datetime").date(2026, 4, 12),
            schedule_entries=[
                ScheduleEntry(
                    entry_type=EntryType.EVENING_ACTIVITY,
                    name="Camp Quiz",
                    start_time="19:00",
                    end_time="20:00",
                    responsible=[leader],
                )
            ],
        )
        camp = Camp(
            name="TestCamp",
            camp_place=CampPlace("TestPlace"),
            participants=[leader],
            schedule=Schedule(days=[day]),
        )

        fill_schedule_sheet(camp, ws)

        matching_cells = [
            cell
            for row in ws.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and "Camp Quiz" in cell.value
        ]

        assert matching_cells
        assert any(cell.font and cell.font.bold for cell in matching_cells)
        assert all("Ali" not in (cell.value or "") for cell in matching_cells)
