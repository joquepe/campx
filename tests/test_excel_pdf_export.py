from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from campx.excel.pdf_export import (
    export_schedule_pdfs,
    _blank_unformatted_cells_for_pdf,
)
from campx.model.camp import Camp
from campx.model.camp_place import CampPlace
from campx.model.day import Day
from campx.model.enums import EntryType, ParticipantType
from campx.model.participant import Participant
from campx.model.schedule import Schedule
from campx.model.schedule_entry import ScheduleEntry


def _build_camp() -> Camp:
    leader_1 = Participant(
        participant_id=1,
        first_name="Alice",
        last_name="Andersson",
        gender="F",
        birthday=date(1995, 1, 1),
        participant_type=ParticipantType.LEADER,
        nick_name="Ali",
        first_name_initials="A",
        last_name_initials="A",
    )
    leader_2 = Participant(
        participant_id=2,
        first_name="Bo",
        last_name="Berg",
        gender="M",
        birthday=date(1994, 1, 1),
        participant_type=ParticipantType.LEADER,
        nick_name="Bo",
        first_name_initials="B",
        last_name_initials="B",
    )
    confirmee = Participant(
        participant_id=3,
        first_name="Cecilia",
        last_name="C",
        gender="F",
        birthday=date(2010, 1, 1),
        participant_type=ParticipantType.CONFIRMEE,
        nick_name="Ce",
    )
    entry = ScheduleEntry(
        entry_type=EntryType.WAKE_UP,
        name="Morning wake-up",
        start_time="08:00",
        end_time="08:15",
        responsible=[leader_1, leader_2],
    )
    day = Day(date=date(2026, 4, 12), schedule_entries=[entry])
    return Camp(
        name="TestCamp",
        camp_place=CampPlace("TestPlace"),
        participants=[leader_1, leader_2, confirmee],
        schedule=Schedule(days=[day]),
    )


def test_export_schedule_pdfs_skips_on_non_macos(tmp_path, monkeypatch):
    camp = _build_camp()
    monkeypatch.setattr("campx.excel.pdf_export.sys.platform", "linux")

    pdfs = export_schedule_pdfs(camp, output_dir=tmp_path)

    assert pdfs == []


def test_export_schedule_pdfs_skips_when_osascript_missing(tmp_path, monkeypatch):
    camp = _build_camp()
    monkeypatch.setattr("campx.excel.pdf_export.sys.platform", "darwin")
    monkeypatch.setattr("campx.excel.pdf_export.shutil.which", lambda _: None)

    pdfs = export_schedule_pdfs(camp, output_dir=tmp_path)

    assert pdfs == []


def test_export_schedule_pdfs_creates_main_and_leader_pdfs(tmp_path, monkeypatch):
    camp = _build_camp()
    monkeypatch.setattr("campx.excel.pdf_export.sys.platform", "darwin")
    monkeypatch.setattr(
        "campx.excel.pdf_export.shutil.which", lambda command: "/usr/bin/osascript"
    )

    def _fake_run(command, check, capture_output, text):
        assert command[0] == "/usr/bin/osascript"
        pdf_path = Path(command[-1])
        pdf_path.touch()

    monkeypatch.setattr("campx.excel.pdf_export.subprocess.run", _fake_run)

    pdfs = export_schedule_pdfs(camp, output_dir=tmp_path)

    pdf_names = sorted(path.name for path in pdfs)
    assert pdf_names == [
        "TestCamp_Ali.pdf",
        "TestCamp_Bo.pdf",
        "TestCamp_Schema.pdf",
    ]


def test_export_schedule_pdfs_uses_absolute_paths_for_osascript(monkeypatch):
    camp = _build_camp()
    monkeypatch.setattr("campx.excel.pdf_export.sys.platform", "darwin")
    monkeypatch.setattr(
        "campx.excel.pdf_export.shutil.which", lambda command: "/usr/bin/osascript"
    )

    captured_args: list[list[str]] = []

    def _fake_run(command, check, capture_output, text):
        captured_args.append(command)
        Path(command[-1]).touch()

    monkeypatch.setattr("campx.excel.pdf_export.subprocess.run", _fake_run)

    pdfs = export_schedule_pdfs(camp)

    assert pdfs
    assert captured_args
    for command in captured_args:
        assert command[3] == "--"
        assert Path(command[4]).is_absolute()
        assert Path(command[5]).is_absolute()


def test_blank_unformatted_cells_for_pdf_only_affects_default_empty_cells():
    workbook = Workbook()
    worksheet = workbook.active

    untouched_empty = worksheet.cell(row=2, column=2)
    styled_empty = worksheet.cell(row=2, column=3)
    styled_empty.fill = PatternFill(
        start_color="00FF0000", end_color="00FF0000", fill_type="solid"
    )
    worksheet.cell(row=1, column=1, value="Header")

    _blank_unformatted_cells_for_pdf(workbook)

    assert worksheet.print_options.gridLines is False
    assert untouched_empty.fill.fill_type == "solid"
    assert getattr(untouched_empty.fill.start_color, "rgb", None) == "00FFFFFF"
    assert getattr(styled_empty.fill.start_color, "rgb", None) == "00FF0000"
