from datetime import date
from openpyxl import Workbook
from campx.excel.excel_generation import (
    add_participant_schedule_sheets,
    add_validation_errors_sheet,
)
from campx.model.camp import Camp
from campx.model.camp_place import CampPlace
from campx.model.day import Day
from campx.model.schedule import Schedule
from campx.model.schedule_entry import ScheduleEntry
from campx.model.enums import EntryType, ParticipantType
from campx.model.participant import Participant


def test_add_participant_schedule_sheets_creates_nickname_sheets():
    leader = Participant(
        participant_id=1,
        first_name="Alice",
        last_name="Andersson",
        gender="F",
        birthday=date(1995, 1, 1),
        participant_type=ParticipantType.LEADER,
        nick_name="Ali",
    )
    confirmand = Participant(
        participant_id=2,
        first_name="Bertil",
        last_name="Berg",
        gender="M",
        birthday=date(2010, 6, 15),
        participant_type=ParticipantType.CONFIRMEE,
        nick_name="Bert",
    )
    entry = ScheduleEntry(
        entry_type=EntryType.WAKE_UP,
        name="Morning wake-up",
        start_time="08:00",
        end_time="08:15",
        responsible=[leader],
    )
    day = Day(date=date(2026, 4, 12), schedule_entries=[entry])
    camp = Camp(
        name="TestCamp",
        camp_place=CampPlace("TestPlace"),
        participants=[leader, confirmand],
        schedule=Schedule(days=[day]),
    )

    wb = Workbook()
    add_participant_schedule_sheets(camp, wb)

    assert "Ali" in wb.sheetnames
    assert "Bert" in wb.sheetnames


def test_participant_schedule_sheet_highlights_target_name():
    leader = Participant(
        participant_id=1,
        first_name="Alice",
        last_name="Andersson",
        gender="F",
        birthday=date(1995, 1, 1),
        participant_type=ParticipantType.LEADER,
        nick_name="Ali",
    )
    entry = ScheduleEntry(
        entry_type=EntryType.WAKE_UP,
        name="Morning wake-up",
        start_time="08:00",
        end_time="08:15",
        responsible=[leader],
    )
    day = Day(date=date(2026, 4, 12), schedule_entries=[entry])
    camp = Camp(
        name="TestCamp",
        camp_place=CampPlace("TestPlace"),
        participants=[leader],
        schedule=Schedule(days=[day]),
    )

    wb = Workbook()
    add_participant_schedule_sheets(camp, wb)
    sheet = wb["Ali"]

    assert any(cell.value == "Participant: Ali" for cell in sheet[1])

    highlighted = [
        cell
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str)
        and "Ali" in cell.value
        and cell.font is not None
        and cell.font.color is not None
        and getattr(cell.font.color, "rgb", None) in {"FF0000", "00FF0000"}
    ]

    assert highlighted


def test_validation_errors_sheet_lists_errors():
    leader = Participant(
        participant_id=1,
        first_name="Alice",
        last_name="Andersson",
        gender="F",
        birthday=date(1995, 1, 1),
        participant_type=ParticipantType.LEADER,
        nick_name="Ali",
    )
    day = Day(date=date(2026, 4, 12), schedule_entries=[])
    camp = Camp(
        name="TestCamp",
        camp_place=CampPlace("TestPlace"),
        participants=[leader],
        schedule=Schedule(days=[day]),
    )

    wb = Workbook()
    add_validation_errors_sheet(camp, wb)

    assert "Valideringsfel" in wb.sheetnames
    sheet = wb["Valideringsfel"]
    assert sheet[2][0].value == "Severity"
    assert sheet[2][1].value == "Error"
    assert any(
        sheet.cell(row=row, column=1).value == "MAJOR"
        and isinstance(sheet.cell(row=row, column=2).value, str)
        and "Leader Alice Andersson has no responsibilities at all"
        in sheet.cell(row=row, column=2).value
        for row in range(3, sheet.max_row + 1)
    )
