from datetime import date
from openpyxl import Workbook
from campx.excel.metrics import (
    _calculate_entry_assignment_metrics,
    _calculate_responsibility_spacing,
    fill_metrics_sheet,
)
from campx.model.camp import Camp
from campx.model.camp_place import CampPlace
from campx.model.day import Day
from campx.model.enums import EntryType, LeaderYearType, ParticipantType
from campx.model.participant import Participant
from campx.model.schedule import Schedule
from campx.model.schedule_entry import ScheduleEntry


def test_fill_metrics_sheet_generates_summary_labels():
    leader_a = Participant(
        participant_id=1,
        first_name="Alice",
        last_name="Andersson",
        gender="F",
        birthday=date(1995, 1, 1),
        participant_type=ParticipantType.LEADER,
    )
    leader_b = Participant(
        participant_id=2,
        first_name="Bo",
        last_name="Bengtsson",
        gender="M",
        birthday=date(1996, 1, 1),
        participant_type=ParticipantType.LEADER,
    )
    entry = ScheduleEntry(
        entry_type=EntryType.WAKE_UP,
        name="Wake up",
        start_time="08:00",
        end_time="08:15",
        responsible=[leader_a],
    )
    day = Day(date=date(2026, 4, 12), schedule_entries=[entry])
    camp = Camp(
        name="TestCamp",
        camp_place=CampPlace("TestPlace"),
        participants=[leader_a, leader_b],
        schedule=Schedule(days=[day]),
    )

    wb = Workbook()
    ws = wb.active
    fill_metrics_sheet(camp, ws)

    assert ws["A1"].value == "Schedule Quality Metrics"
    assert any(
        ws.cell(row=row, column=1).value == "Total leaders" for row in range(1, 20)
    )
    assert any(
        ws.cell(row=row, column=5).value == "Constrained entries"
        for row in range(1, 20)
    )
    assert any(
        ws.cell(row=row, column=1).value
        == "Average gap between selected responsibilities (h)"
        for row in range(1, ws.max_row + 1)
    )
    assert any(
        ws.cell(row=row, column=1).value == "Entries without experienced cover"
        for row in range(1, ws.max_row + 1)
    )
    assert any(
        ws.cell(row=row, column=10).value == "Entry allocation profile"
        for row in range(1, 20)
    )


def test_calculate_responsibility_spacing_supports_entry_type_filtering():
    leader = Participant(
        participant_id=1,
        first_name="Alice",
        last_name="Andersson",
        gender="F",
        birthday=date(1995, 1, 1),
        participant_type=ParticipantType.LEADER,
    )
    other_leader = Participant(
        participant_id=2,
        first_name="Bo",
        last_name="Bengtsson",
        gender="M",
        birthday=date(1996, 1, 1),
        participant_type=ParticipantType.LEADER,
    )

    wake_up = ScheduleEntry(
        entry_type=EntryType.WAKE_UP,
        name="Wake up",
        start_time="08:00",
        end_time="08:15",
        responsible=[leader],
    )
    morning_prayer = ScheduleEntry(
        entry_type=EntryType.MORNING_PRAYER,
        name="Morning prayer",
        start_time="09:00",
        end_time="09:30",
        responsible=[leader],
    )
    put_to_bed = ScheduleEntry(
        entry_type=EntryType.PUT_TO_BED,
        name="Put to bed",
        start_time="22:30",
        end_time="23:00",
        responsible=[leader],
    )
    day_host = ScheduleEntry(
        entry_type=EntryType.DAY_HOST,
        name="Day host",
        start_time="00:00",
        end_time="00:00",
        responsible=[leader],
    )

    day = Day(
        date=date(2026, 4, 12),
        schedule_entries=[wake_up, morning_prayer, put_to_bed, day_host],
    )
    camp = Camp(
        name="TestCamp",
        camp_place=CampPlace("TestPlace"),
        participants=[leader, other_leader],
        schedule=Schedule(days=[day]),
    )

    filtered = _calculate_responsibility_spacing(
        camp,
        included_entry_types={EntryType.WAKE_UP, EntryType.MORNING_PRAYER},
        tight_gap_minutes=120,
    )
    filtered_metrics = next(
        item
        for item in filtered["leader_spacing"]
        if item["leader_name"] == leader.full_name
    )

    assert filtered_metrics["gap_count"] == 1
    assert filtered_metrics["tight_gap_count"] == 1
    assert round(filtered_metrics["avg_gap_hours"], 2) == 0.75

    unfiltered = _calculate_responsibility_spacing(camp, tight_gap_minutes=120)
    unfiltered_metrics = next(
        item
        for item in unfiltered["leader_spacing"]
        if item["leader_name"] == leader.full_name
    )

    assert unfiltered_metrics["gap_count"] == 2
    assert unfiltered_metrics["tight_gap_count"] == 1


def test_calculate_entry_assignment_metrics_for_required_entries_only():
    first_year_leader = Participant(
        participant_id=1,
        first_name="Alice",
        last_name="Andersson",
        gender="F",
        birthday=date(2000, 1, 1),
        participant_type=ParticipantType.LEADER,
        leader_year_type=LeaderYearType.FIRST_YEAR,
    )
    returning_leader = Participant(
        participant_id=2,
        first_name="Bo",
        last_name="Bengtsson",
        gender="M",
        birthday=date(1998, 1, 1),
        participant_type=ParticipantType.LEADER,
        leader_year_type=LeaderYearType.SECOND_YEAR,
    )
    management = Participant(
        participant_id=3,
        first_name="Cia",
        last_name="Chef",
        gender="F",
        birthday=date(1990, 1, 1),
        participant_type=ParticipantType.CAMP_MANAGEMENT,
    )

    wake_up = ScheduleEntry(
        entry_type=EntryType.WAKE_UP,
        name="Wake up",
        start_time="08:00",
        end_time="08:15",
        responsible=[first_year_leader, management],
    )
    evening_activity = ScheduleEntry(
        entry_type=EntryType.EVENING_ACTIVITY,
        name="Evening activity",
        start_time="19:00",
        end_time="20:00",
        responsible=[returning_leader],
    )
    dinner = ScheduleEntry(
        entry_type=EntryType.DINNER,
        name="Dinner",
        start_time="18:00",
        end_time="19:00",
        responsible=[],
    )
    day = Day(
        date=date(2026, 4, 12),
        schedule_entries=[wake_up, evening_activity, dinner],
    )
    camp = Camp(
        name="TestCamp",
        camp_place=CampPlace("TestPlace"),
        participants=[first_year_leader, returning_leader, management],
        schedule=Schedule(days=[day]),
    )

    metrics = _calculate_entry_assignment_metrics(camp)

    assert len(metrics) == 2
    wake_up_metrics = next(
        item for item in metrics if item["entry"].entry_type == EntryType.WAKE_UP
    )
    assert wake_up_metrics["responsible_count"] == 2
    assert wake_up_metrics["first_year_count"] == 1
    assert wake_up_metrics["returning_leader_count"] == 0
    assert wake_up_metrics["management_count"] == 1
    assert wake_up_metrics["gender_mix"] == "F:2"
    assert wake_up_metrics["has_experienced_cover"] is True


def test_fill_metrics_sheet_includes_entry_allocation_values():
    first_year_leader = Participant(
        participant_id=1,
        first_name="Alice",
        last_name="Andersson",
        gender="F",
        birthday=date(2000, 1, 1),
        participant_type=ParticipantType.LEADER,
        leader_year_type=LeaderYearType.FIRST_YEAR,
    )
    second_year_leader = Participant(
        participant_id=2,
        first_name="Bo",
        last_name="Bengtsson",
        gender="M",
        birthday=date(1998, 1, 1),
        participant_type=ParticipantType.LEADER,
        leader_year_type=LeaderYearType.SECOND_YEAR,
    )

    wake_up = ScheduleEntry(
        entry_type=EntryType.WAKE_UP,
        name="Wake up",
        start_time="08:00",
        end_time="08:15",
        responsible=[first_year_leader],
    )
    evening_activity = ScheduleEntry(
        entry_type=EntryType.EVENING_ACTIVITY,
        name="Evening activity",
        start_time="19:00",
        end_time="20:00",
        responsible=[first_year_leader, second_year_leader],
    )
    day = Day(
        date=date(2026, 4, 12),
        schedule_entries=[wake_up, evening_activity],
    )
    camp = Camp(
        name="TestCamp",
        camp_place=CampPlace("TestPlace"),
        participants=[first_year_leader, second_year_leader],
        schedule=Schedule(days=[day]),
    )

    wb = Workbook()
    ws = wb.active
    fill_metrics_sheet(camp, ws)

    assert any(
        ws.cell(row=row, column=1).value == "Entries without experienced cover"
        and ws.cell(row=row, column=2).value == 1
        for row in range(1, ws.max_row + 1)
    )
    assert any(
        ws.cell(row=row, column=11).value == EntryType.WAKE_UP.value.short
        and ws.cell(row=row, column=13).value == 1
        and ws.cell(row=row, column=14).value == 0
        and ws.cell(row=row, column=17).value == "F:1"
        for row in range(1, ws.max_row + 1)
    )
