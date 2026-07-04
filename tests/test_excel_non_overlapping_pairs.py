from datetime import date

from openpyxl import Workbook

from campx.excel.excel_generation import add_non_overlapping_pairs_sheet
from campx.excel.non_overlapping_pairs import fill_non_overlapping_pairs_sheet
from campx.model.camp import Camp
from campx.model.camp_place import CampPlace
from campx.model.day import Day
from campx.model.enums import EntryType, ParticipantType
from campx.model.participant import Participant
from campx.model.schedule import Schedule
from campx.model.schedule_entry import ScheduleEntry


def _leader(participant_id: int, first_name: str, last_name: str) -> Participant:
    return Participant(
        participant_id=participant_id,
        first_name=first_name,
        last_name=last_name,
        gender="F",
        birthday=date(1995, 1, 1),
        participant_type=ParticipantType.LEADER,
    )


def test_fill_non_overlapping_pairs_sheet_lists_only_valid_pairs():
    alice = _leader(1, "Alice", "Andersson")
    bo = _leader(2, "Bo", "Bengtsson")
    carl = _leader(3, "Carl", "Carlsson")

    day_1 = Day(
        date=date(2026, 7, 1),
        schedule_entries=[
            ScheduleEntry(
                entry_type=EntryType.DAY_OFF,
                name="",
                start_time=None,
                end_time=None,
                responsible=[alice],
            ),
            ScheduleEntry(
                entry_type=EntryType.SLEEP_IN,
                name="",
                start_time=None,
                end_time=None,
                responsible=[bo],
            ),
        ],
    )
    day_2 = Day(
        date=date(2026, 7, 2),
        schedule_entries=[
            ScheduleEntry(
                entry_type=EntryType.DAY_OFF,
                name="",
                start_time=None,
                end_time=None,
                responsible=[bo],
            ),
            ScheduleEntry(
                entry_type=EntryType.SLEEP_IN,
                name="",
                start_time=None,
                end_time=None,
                responsible=[carl],
            ),
        ],
    )

    camp = Camp(
        name="TestCamp",
        camp_place=CampPlace("TestPlace"),
        participants=[alice, bo, carl],
        schedule=Schedule(days=[day_1, day_2]),
    )

    wb = Workbook()
    ws = wb.active
    fill_non_overlapping_pairs_sheet(camp, ws)

    assert ws.cell(row=1, column=1).value == "Leader 1"
    assert ws.cell(row=1, column=2).value == "Leader 2"

    listed_pairs = {
        (ws.cell(row=row, column=1).value, ws.cell(row=row, column=2).value)
        for row in range(2, ws.max_row + 1)
        if ws.cell(row=row, column=1).value and ws.cell(row=row, column=2).value
    }

    assert listed_pairs == {
        ("Alice Andersson", "Carl Carlsson"),
        ("Carl Carlsson", "Alice Andersson"),
    }


def test_add_non_overlapping_pairs_sheet_creates_named_sheet():
    alice = _leader(1, "Alice", "Andersson")
    bo = _leader(2, "Bo", "Bengtsson")
    camp = Camp(
        name="TestCamp",
        camp_place=CampPlace("TestPlace"),
        participants=[alice, bo],
        schedule=Schedule(days=[]),
    )

    wb = Workbook()
    add_non_overlapping_pairs_sheet(camp, wb)

    assert "Ej överlapp" in wb.sheetnames
